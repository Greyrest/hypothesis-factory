"""Парсер отчётов института по хвостам (xlsx) в структурированный словарь.

Устойчив к неполным данным: ячейки #REF!, пропуски, 1 или 2 потока хвостов
(породные / пирротиновые), разный набор классов крупности. Перенос
parse_tailings.py прототипа без изменения алгоритма.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from hf_domains.mining_flotation.constants import FORM_ALIASES, RECOVERABLE_FORMS

SIZE_RE = re.compile(r"^\s*([+-]\s*\d+(?:\s*[+-]\s*\d+)?)\s*(?:мкм)?\s*$", re.I)


def _num(v):
    """Число или None. '#REF!' и прочий мусор -> None."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def _text(v):
    if isinstance(v, str):
        return v.strip()
    return None


def canon_class(label: str) -> str | None:
    """' -20 + 10 мкм' -> '-20+10'."""
    m = SIZE_RE.match(label.replace("мкм", " "))
    if not m:
        return None
    return re.sub(r"\s+", "", m.group(1))


def canon_form(label: str) -> str | None:
    low = label.lower()
    for key, canon in FORM_ALIASES:
        if low.startswith(key):
            return canon
    return None


def _row_cells(row):
    """[(col_idx, value), ...] только непустые."""
    out = []
    for c in row:
        if c.value is not None and _text(c.value) != "":
            out.append((c.column, c.value))
    return out


def _header_colmap(cells):
    """По заголовку блока определяем, в каких колонках лежат метрики."""
    colmap = {}
    for col, v in cells:
        t = _text(v)
        if not t:
            continue
        low = t.lower()
        if "доля класса" in low:
            colmap["share"] = col
        elif "доля" in low and "28" in low:
            colmap["ni_share"] = col
        elif "28" in low and "т" in low:
            colmap["ni_t"] = col
        elif "доля" in low and "29" in low:
            colmap["cu_share"] = col
        elif "29" in low and "т" in low:
            colmap["cu_t"] = col
    return colmap


def _pick(cells, colmap, key):
    col = colmap.get(key)
    if col is None:
        return None
    for c, v in cells:
        if c == col:
            return _num(v)
    return None


def parse_workbook(path: str | Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Итог"] if "Итог" in wb.sheetnames else wb.worksheets[0]

    result = {
        "source_file": str(path),
        "plant": Path(path).stem.replace("Хвосты", "").replace("_2", "").strip(),
        "feed": {},
        "tailings_fact": None,
        "streams": [],
        "warnings": [],
    }

    stream = None          # текущий поток хвостов
    mode = None            # None | 'class_table' | 'form_block'
    colmap = {}
    cur_class = None       # канонич. имя класса для form_block

    for row in ws.iter_rows():
        cells = _row_cells(row)
        if not cells:
            continue
        label = _text(cells[0][1]) or ""
        nums = [(c, _num(v)) for c, v in cells[1:]]
        numvals = [v for _, v in nums if v is not None]
        low = label.lower()

        # --- шапка: питание фабрики ---
        if low.startswith("шихта руд") and len(numvals) >= 3:
            result["feed"]["ore"] = {"smt": numvals[0], "ni_pct": numvals[1],
                                     "ni_t": numvals[2],
                                     "cu_pct": numvals[3] if len(numvals) > 3 else None,
                                     "cu_t": numvals[4] if len(numvals) > 4 else None}
            continue
        if low.startswith("отвальные хвосты") and len(numvals) >= 3 and result["tailings_fact"] is None:
            result["tailings_fact"] = {"smt": numvals[0], "ni_pct": numvals[1],
                                       "ni_t": numvals[2],
                                       "cu_pct": numvals[3] if len(numvals) > 3 else None,
                                       "cu_t": numvals[4] if len(numvals) > 4 else None}
            continue

        # --- новый поток хвостов (нужны тоннаж + содержания) ---
        if low.startswith("хвосты") and len(numvals) >= 4:
            stream = {
                "name": label,
                "smt": numvals[0],
                "ni_pct": numvals[1], "ni_t": numvals[2],
                "cu_pct": numvals[3], "cu_t": numvals[4] if len(numvals) > 4 else None,
                "size_classes": [],
                "totals": {},
            }
            result["streams"].append(stream)
            mode = None
            continue

        # --- заголовок таблицы распределения по классам ---
        if low.startswith("класс крупности"):
            mode = "class_table"
            colmap = _header_colmap(cells)
            continue

        # --- заголовок блока минеральных форм: '+71 мкм | Доля потерь...' ---
        cc = canon_class(label) if "мкм" in low or SIZE_RE.match(label) else None
        header_texts = " ".join(_text(v) or "" for _, v in cells[1:]).lower()
        if cc and "доля потерь" in header_texts:
            mode = "form_block"
            cur_class = cc
            colmap = _header_colmap(cells)
            if stream is not None and not any(s["cls"] == cc for s in stream["size_classes"]):
                stream["size_classes"].append({"cls": cc, "forms": {}})
            continue

        if stream is None:
            continue

        # --- строки таблицы классов ---
        if mode == "class_table":
            if low.startswith("итого"):
                mode = None
                continue
            cc2 = canon_class(label)
            if cc2:
                entry = next((s for s in stream["size_classes"] if s["cls"] == cc2), None)
                if entry is None:
                    entry = {"cls": cc2, "forms": {}}
                    stream["size_classes"].append(entry)
                entry.update({
                    "share_pct": _pick(cells, colmap, "share"),
                    "ni_share_pct": _pick(cells, colmap, "ni_share"),
                    "ni_t": _pick(cells, colmap, "ni_t"),
                    "cu_share_pct": _pick(cells, colmap, "cu_share"),
                    "cu_t": _pick(cells, colmap, "cu_t"),
                })
            continue

        # --- строки блока форм ---
        if mode == "form_block":
            entry = next((s for s in stream["size_classes"] if s["cls"] == cur_class), None)
            if entry is None:
                continue
            if low.startswith("итого"):
                continue
            if low.startswith("извлекаемый") or low.startswith("не извлекаемый"):
                key = "recoverable" if low.startswith("извлекаемый") else "non_recoverable"
                entry[key] = {
                    "ni_share_pct": _pick(cells, colmap, "ni_share"),
                    "ni_t": _pick(cells, colmap, "ni_t"),
                    "cu_share_pct": _pick(cells, colmap, "cu_share"),
                    "cu_t": _pick(cells, colmap, "cu_t"),
                }
                continue
            form = canon_form(label)
            if form:
                entry["forms"][form] = {
                    "ni_share_pct": _pick(cells, colmap, "ni_share"),
                    "ni_t": _pick(cells, colmap, "ni_t"),
                    "cu_share_pct": _pick(cells, colmap, "cu_share"),
                    "cu_t": _pick(cells, colmap, "cu_t"),
                }
            continue

    _postprocess(result)
    return result


def _postprocess(result: dict):
    """Достраиваем извлекаемый металл там, где в отчёте #REF!/пусто, и итоги."""
    # сводный блок (сумма остальных потоков) помечаем, чтобы не задваивать
    for i, stream in enumerate(result["streams"]):
        prev_smt = sum(s["smt"] or 0 for s in result["streams"][:i])
        stream["aggregate"] = bool(
            i >= 2 and prev_smt and abs((stream["smt"] or 0) - prev_smt) / prev_smt < 0.01)

    for stream in result["streams"]:
        tot = {"ni_t": 0.0, "cu_t": 0.0, "rec_ni_t": 0.0, "rec_cu_t": 0.0}
        for entry in stream["size_classes"]:
            # если 'Извлекаемый металл' битый — считаем сами из форм
            for el in ("ni", "cu"):
                rec = entry.get("recoverable", {}) or {}
                if rec.get(f"{el}_t") is None and entry["forms"]:
                    s = 0.0
                    have = False
                    for form, vals in entry["forms"].items():
                        v = vals.get(f"{el}_t")
                        if v is not None and form in RECOVERABLE_FORMS[el]:
                            s += v
                            have = True
                    if have:
                        entry.setdefault("recoverable", {})[f"{el}_t"] = round(s, 2)
                        entry["recoverable"].setdefault("estimated", True)
                        result["warnings"].append(
                            f"{stream['name']} / {entry['cls']}: извлекаемый {el} "
                            f"восстановлен из форм (в отчёте нет данных)")
            # если тонны потерь класса нет, но есть формы — суммируем
            for el in ("ni", "cu"):
                if entry.get(f"{el}_t") is None and entry["forms"]:
                    vals = [v.get(f"{el}_t") for v in entry["forms"].values()
                            if v.get(f"{el}_t") is not None]
                    if vals:
                        entry[f"{el}_t"] = round(sum(vals), 2)
            tot["ni_t"] += entry.get("ni_t") or 0.0
            tot["cu_t"] += entry.get("cu_t") or 0.0
            rec = entry.get("recoverable", {}) or {}
            tot["rec_ni_t"] += rec.get("ni_t") or 0.0
            tot["rec_cu_t"] += rec.get("cu_t") or 0.0
        stream["totals"] = {
            "ni_t": round(tot["ni_t"], 1),
            "cu_t": round(tot["cu_t"], 1),
            "recoverable_ni_t": round(tot["rec_ni_t"], 1),
            "recoverable_cu_t": round(tot["rec_cu_t"], 1),
        }
