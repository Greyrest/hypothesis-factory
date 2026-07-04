"""Парсер отчётов института по хвостам (xlsx) в структурированный JSON.

Устойчив к неполным данным: ячейки #REF!, пропуски, 1 или 2 потока хвостов
(породные / пирротиновые), разный набор классов крупности.

Компоненты потерь («Элемент N») извлекаются из самого отчёта: конвейер не
завязан на конкретные металлы. Для известных атомных номеров id компонента —
символ элемента (28 -> ni, 29 -> cu), для прочих — el<N>. Все метрики дальше
ключуются id компонента (поля вида <id>_t / <id>_pct и словари по id).
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

# Канонические формы нахождения металла в хвостах
FORM_ALIASES = [
    ("раскрытый", "Раскрытый Pnt/Cp"),
    ("закрытый", "Закрытый Pnt/Cp"),
    ("примесь в пирротине", "Примесь в пирротине"),
    ("силикатная", "Силикатная форма/Валлериит"),
    ("пирит", "Пирит/Другие"),
    ("миллерит", "Миллерит"),
    ("потери (расписать", "Потери (расписать)"),
    ("свободный слот", "Свободный слот"),
]

# Извлекаемые формы (по справке института): базово — раскрытые/закрытые
# сульфиды; для отдельных элементов добавляются свои извлекаемые минералы
# (эл.28/Ni — миллерит). Ключ — id компонента.
BASE_RECOVERABLE_FORMS = {"Раскрытый Pnt/Cp", "Закрытый Pnt/Cp"}
EXTRA_RECOVERABLE_FORMS = {"ni": {"Миллерит"}}

# известные атомные номера -> (id, символ); прочие получают id el<N>
ELEMENT_ALIASES = {28: ("ni", "Ni"), 29: ("cu", "Cu")}
# формат отчёта института по умолчанию, если в файле не нашлось упоминаний
DEFAULT_ELEMENTS = [28, 29]

ELEMENT_RE = re.compile(r"(?:элемент|эл)\.?\s*(\d+)", re.I)

SIZE_RE = re.compile(r"^\s*([+-]\s*\d+(?:\s*[+-]\s*\d+)?)\s*(?:мкм)?\s*$", re.I)


def _component(num: int) -> dict:
    """Описание компонента потерь по номеру элемента из отчёта."""
    alias = ELEMENT_ALIASES.get(num)
    cid = alias[0] if alias else f"el{num}"
    label = f"Элемент {num} ({alias[1]})" if alias else f"Элемент {num}"
    forms = BASE_RECOVERABLE_FORMS | EXTRA_RECOVERABLE_FORMS.get(cid, set())
    return {"id": cid, "num": num, "label": label, "unit": "т",
            "recoverable_forms": sorted(forms)}


def _discover_components(ws) -> list[dict]:
    """Компоненты в порядке первого упоминания «Элемент N» / «эл.N» в отчёте."""
    nums: list[int] = []
    for row in ws.iter_rows():
        for c in row:
            t = _text(c.value)
            if not t:
                continue
            for m in ELEMENT_RE.finditer(t.lower()):
                n = int(m.group(1))
                if n not in nums:
                    nums.append(n)
    return [_component(n) for n in (nums or DEFAULT_ELEMENTS)]


def _num(v):
    """Число или None. '#REF!' и прочий мусор -> None."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def _text(v):
    if isinstance(v, str):
        return v.strip()
    return None


def canon_class(label: str) -> str | None:
    """' -20 + 10 мкм' -> '-20+10'."""
    m = SIZE_RE.match(label.replace("мкм", " "))
    if not m:
        return None
    return re.sub(r"\s+", "", m.group(1))


def canon_form(label: str) -> str | None:
    low = label.lower()
    for key, canon in FORM_ALIASES:
        if low.startswith(key):
            return canon
    return None


def _row_cells(row):
    """[(col_idx, value), ...] только непустые."""
    out = []
    for c in row:
        if c.value is not None and _text(c.value) != "":
            out.append((c.column, c.value))
    return out


def _header_colmap(cells, components):
    """По заголовку блока определяем, в каких колонках лежат метрики."""
    colmap = {}
    for col, v in cells:
        t = _text(v)
        if not t:
            continue
        low = t.lower()
        if "доля класса" in low:
            colmap["share"] = col
            continue
        for comp in components:
            num = str(comp["num"])
            if num not in low:
                continue
            if "доля" in low:
                colmap[f"{comp['id']}_share"] = col
            elif "т" in low:
                colmap[f"{comp['id']}_t"] = col
            break
    return colmap


def _pick(cells, colmap, key):
    col = colmap.get(key)
    if col is None:
        return None
    for c, v in cells:
        if c == col:
            return _num(v)
    return None


def parse_workbook(src) -> dict:
    """src — путь или файловый объект (BytesIO): диск не обязателен."""
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Итог"] if "Итог" in wb.sheetnames else wb.worksheets[0]

    components = _discover_components(ws)
    comp_ids = [c["id"] for c in components]

    def _row_metrics(numvals):
        """Позиционная строка «smt, затем пары (%, т) на компонент»."""
        out = {"smt": numvals[0]}
        for i, cid in enumerate(comp_ids):
            pct = numvals[1 + 2 * i] if len(numvals) > 1 + 2 * i else None
            t = numvals[2 + 2 * i] if len(numvals) > 2 + 2 * i else None
            out[f"{cid}_pct"] = pct
            out[f"{cid}_t"] = t
        return out

    is_path = isinstance(src, (str, Path))
    stem = Path(src).stem if is_path else ""
    result = {
        "source_file": str(src) if is_path else "upload",
        "plant": stem.replace("Хвосты", "").replace("_2", "").strip(),
        "components": components,
        "feed": {},
        "tailings_fact": None,
        "streams": [],
        "warnings": [],
    }

    stream = None          # текущий поток хвостов
    mode = None            # None | 'class_table' | 'form_block'
    colmap = {}
    cur_class = None       # канонич. имя класса для form_block

    for row in ws.iter_rows():
        cells = _row_cells(row)
        if not cells:
            continue
        label = _text(cells[0][1]) or ""
        nums = [(c, _num(v)) for c, v in cells[1:]]
        numvals = [v for _, v in nums if v is not None]
        low = label.lower()

        # --- шапка: питание фабрики ---
        if low.startswith("шихта руд") and len(numvals) >= 3:
            result["feed"]["ore"] = _row_metrics(numvals)
            continue
        if low.startswith("итого") and result["feed"].get("ore") and not result["feed"].get("total") and len(numvals) >= 3 and stream is None and mode is None:
            result["feed"]["total"] = _row_metrics(numvals)
            continue
        if low.startswith("отвальные хвосты") and len(numvals) >= 3 and result["tailings_fact"] is None:
            result["tailings_fact"] = _row_metrics(numvals)
            continue

        # --- новый поток хвостов (нужны тоннаж + содержания) ---
        if low.startswith("хвосты") and len(numvals) >= max(3, 2 * len(comp_ids)):
            stream = {
                "name": label,
                **_row_metrics(numvals),
                "size_classes": [],
                "totals": {},
            }
            result["streams"].append(stream)
            mode = None
            continue

        # --- заголовок таблицы распределения по классам ---
        if low.startswith("класс крупности"):
            mode = "class_table"
            colmap = _header_colmap(cells, components)
            continue

        # --- заголовок блока минеральных форм: '+71 мкм | Доля потерь...' ---
        cc = canon_class(label) if "мкм" in low or SIZE_RE.match(label) else None
        header_texts = " ".join(_text(v) or "" for _, v in cells[1:]).lower()
        if cc and "доля потерь" in header_texts:
            mode = "form_block"
            cur_class = cc
            colmap = _header_colmap(cells, components)
            if stream is not None and not any(s["cls"] == cc for s in stream["size_classes"]):
                stream["size_classes"].append({"cls": cc, "forms": {}})
            continue

        if stream is None:
            continue

        # --- строки таблицы классов ---
        if mode == "class_table":
            if low.startswith("итого"):
                mode = None
                continue
            cc2 = canon_class(label)
            if cc2:
                entry = next((s for s in stream["size_classes"] if s["cls"] == cc2), None)
                if entry is None:
                    entry = {"cls": cc2, "forms": {}}
                    stream["size_classes"].append(entry)
                entry["share_pct"] = _pick(cells, colmap, "share")
                for cid in comp_ids:
                    entry[f"{cid}_share_pct"] = _pick(cells, colmap, f"{cid}_share")
                    entry[f"{cid}_t"] = _pick(cells, colmap, f"{cid}_t")
            continue

        # --- строки блока форм ---
        if mode == "form_block":
            entry = next((s for s in stream["size_classes"] if s["cls"] == cur_class), None)
            if entry is None:
                continue
            if low.startswith("итого"):
                continue
            metrics = {}
            for cid in comp_ids:
                metrics[f"{cid}_share_pct"] = _pick(cells, colmap, f"{cid}_share")
                metrics[f"{cid}_t"] = _pick(cells, colmap, f"{cid}_t")
            if low.startswith("извлекаемый") or low.startswith("не извлекаемый"):
                key = "recoverable" if low.startswith("извлекаемый") else "non_recoverable"
                entry[key] = metrics
                continue
            form = canon_form(label)
            if form:
                entry["forms"][form] = metrics
            continue

    _postprocess(result)
    return result


def _postprocess(result: dict):
    """Достраиваем извлекаемый металл там, где в отчёте #REF!/пусто, и итоги."""
    components = result["components"]
    # сводный блок (сумма остальных потоков) помечаем, чтобы не задваивать
    for i, stream in enumerate(result["streams"]):
        prev_smt = sum(s["smt"] or 0 for s in result["streams"][:i])
        stream["aggregate"] = bool(
            i >= 2 and prev_smt and abs((stream["smt"] or 0) - prev_smt) / prev_smt < 0.01)

    for stream in result["streams"]:
        tot = {c["id"]: 0.0 for c in components}
        rec_tot = {c["id"]: 0.0 for c in components}
        for entry in stream["size_classes"]:
            # если 'Извлекаемый металл' битый — считаем сами из форм
            for comp in components:
                el = comp["id"]
                rec = entry.get("recoverable", {}) or {}
                if rec.get(f"{el}_t") is None and entry["forms"]:
                    s = 0.0
                    have = False
                    for form, vals in entry["forms"].items():
                        v = vals.get(f"{el}_t")
                        if v is not None and form in comp["recoverable_forms"]:
                            s += v
                            have = True
                    if have:
                        entry.setdefault("recoverable", {})[f"{el}_t"] = round(s, 2)
                        entry["recoverable"].setdefault("estimated", True)
                        result["warnings"].append(
                            f"{stream['name']} / {entry['cls']}: извлекаемый {el} "
                            f"восстановлен из форм (в отчёте нет данных)")
            # если тонны потерь класса нет, но есть формы — суммируем
            for comp in components:
                el = comp["id"]
                if entry.get(f"{el}_t") is None and entry["forms"]:
                    vals = [v.get(f"{el}_t") for v in entry["forms"].values()
                            if v.get(f"{el}_t") is not None]
                    if vals:
                        entry[f"{el}_t"] = round(sum(vals), 2)
            rec = entry.get("recoverable", {}) or {}
            for comp in components:
                el = comp["id"]
                tot[el] += entry.get(f"{el}_t") or 0.0
                rec_tot[el] += rec.get(f"{el}_t") or 0.0
        stream["totals"] = {}
        for comp in components:
            el = comp["id"]
            stream["totals"][f"{el}_t"] = round(tot[el], 1)
            stream["totals"][f"recoverable_{el}_t"] = round(rec_tot[el], 1)


def main():
    if len(sys.argv) < 2:
        print("usage: parse_tailings.py <xlsx> [out.json]")
        sys.exit(1)
    data = parse_workbook(sys.argv[1])
    out = json.dumps(data, ensure_ascii=False, indent=2)
    if len(sys.argv) > 2:
        Path(sys.argv[2]).write_text(out, encoding="utf-8")
    else:
        print(out)


if __name__ == "__main__":
    main()
